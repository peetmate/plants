"""v1.1 apply step — merge sidecar edits back into the source-of-truth workbook.

Follows the project's safe protocol (Plants/Claude/README.md + the dispatch):
  1. work on a LOCAL COPY, never the Drive file directly
  2. openpyxl: write notes into the collection Notes column, keyed by UID (col A);
     culture/formula columns are never touched
  3. LibreOffice headless recalc (openpyxl save drops cached formula values; the
     downstream readers use data_only, so caches must be rebuilt) — forced via
     fullCalcOnLoad + a seeded LO profile with OOXML recalc = Always
  4. verify readback: notes written; formula caches restored (col F key resolves);
     validate_workbook.py clean; sheet set intact
  5. optimistic-lock copy-back: abort if the Drive workbook changed meanwhile
  6. back up the Drive workbook, then atomically replace it
  7. covers -> Photos/covers.json ; log to apply_log ; clear applied sidecar rows

Nothing touches the Drive workbook until every check passes.
"""
import json
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

import config
import db

SOFFICE_CANDIDATES = [
    "soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
]

# LO profile that forces recalculation of OOXML (and ODF) formulas on load
# (0 = Always recalculate; without this, headless load may skip recalc).
_RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""


def find_soffice():
    for c in SOFFICE_CANDIDATES:
        p = shutil.which(c)
        if p:
            return p
        if Path(c).exists():
            return c
    return None


def latest_workbook():
    books = sorted(p for p in config.DATA_DIR.glob("*- Orchids.xlsx") if not p.name.startswith("~"))
    if not books:
        raise FileNotFoundError("no '*- Orchids.xlsx' in DATA_DIR")
    return books[-1]


def _recalc(soffice, src: Path, outdir: Path) -> Path:
    profile = outdir / "loprofile"
    (profile / "user").mkdir(parents=True, exist_ok=True)
    (profile / "user" / "registrymodifications.xcu").write_text(_RECALC_XCU, encoding="utf-8")
    cmd = [
        soffice, "--headless", "--norestore", "--nologo", "--nofirststartwizard",
        f"-env:UserInstallation=file://{profile}",
        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
        "--outdir", str(outdir), str(src),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    out = outdir / src.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice recalc failed:\nstdout: {r.stdout}\nstderr: {r.stderr}")
    return out


def _validator_issues(xlsx: Path, work: Path, tag: str):
    """Run the project validator against `xlsx` in an isolated tree; return the set
    of issue lines it reports (or None if the validator isn't available)."""
    validator = config.DATA_DIR / "validate_workbook.py"
    vocab = config.DATA_DIR.parent / "Claude" / "vocab.json"
    if not validator.exists():
        return None
    vroot = work / ("val_" + tag)
    (vroot / "Collection").mkdir(parents=True, exist_ok=True)
    (vroot / "Claude").mkdir(parents=True, exist_ok=True)
    shutil.copy2(xlsx, vroot / "Collection" / xlsx.name)   # name must match *- Orchids.xlsx
    shutil.copy2(validator, vroot / "Collection" / "validate_workbook.py")
    if vocab.exists():
        shutil.copy2(vocab, vroot / "Claude" / "vocab.json")
    import sys
    r = subprocess.run(
        [sys.executable, "validate_workbook.py"],
        cwd=str(vroot / "Collection"), capture_output=True, text=True, timeout=120,
    )
    return {ln.strip() for ln in (r.stdout or "").splitlines() if ln.strip().startswith("- ")}


def _check_no_regression(orig: Path, rec: Path, work: Path):
    """Fail only if apply INTRODUCED validation issues (the workbook may already
    have pre-existing warnings unrelated to this edit)."""
    base = _validator_issues(orig, work, "base")
    after = _validator_issues(rec, work, "rec")
    if base is None or after is None:
        return
    new = after - base
    if new:
        raise RuntimeError("apply introduced validation issues: " + " | ".join(sorted(new)))


def _verify_readback(rec: Path, note_map, applied_uids):
    chk = openpyxl.load_workbook(rec, data_only=True, read_only=True)
    try:
        if "collection" not in chk.sheetnames:
            raise RuntimeError("recalc output missing 'collection' sheet")
        ws = chk["collection"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        uid_i = hdr.index("UID")
        note_i = hdr.index("Notes")
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if isinstance(r[uid_i], (int, float))]
        by_uid = {int(r[uid_i]): r for r in rows}
        # notes written correctly
        for uid in applied_uids:
            got = by_uid.get(uid, [None] * len(hdr))[note_i]
            if (got or "") != (note_map[uid] or ""):
                raise RuntimeError(f"note readback mismatch for UID {uid}")
        # formula caches restored: col F (index 5) key must resolve for most rows
        fcol = 5
        nonnull = sum(1 for r in rows if r[fcol] not in (None, ""))
        if rows and nonnull / len(rows) < 0.9:
            raise RuntimeError(
                f"formula recalc looks broken: only {nonnull}/{len(rows)} col-F keys resolved"
            )
        return len(rows), nonnull
    finally:
        chk.close()


def _merge_covers(covers):
    path = config.COVERS_JSON
    try:
        cur = json.loads(path.read_text(encoding="utf-8") or "{}")
    except (OSError, ValueError):
        cur = {}
    for uid, photo in covers.items():
        cur[str(uid)] = photo
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cur, indent=1), encoding="utf-8")
    return len(covers)


def _log_and_clear(summary, applied_note_uids, cover_uids):
    with db._conn() as c:
        c.execute(
            "INSERT INTO apply_log(ran_at, summary) VALUES(?,?)",
            (datetime.now(timezone.utc).isoformat(), summary),
        )
        c.executemany("DELETE FROM note_edits WHERE uid=?", [(u,) for u in applied_note_uids])
        c.executemany("DELETE FROM cover_edits WHERE uid=?", [(u,) for u in cover_uids])


def apply_edits():
    """Returns (payload dict, http status)."""
    soffice = find_soffice()
    if not soffice:
        return {"ok": False, "error": "LibreOffice (soffice) not found — install it to apply."}, 503

    notes = db.get_notes()      # {uid: note}
    covers = db.get_covers()    # {uid: photo}
    if not notes and not covers:
        return {"ok": True, "applied_notes": 0, "applied_covers": 0, "note": "nothing to apply"}, 200

    wb_path = latest_workbook()
    start_mtime = wb_path.stat().st_mtime_ns

    work = Path(tempfile.mkdtemp(prefix="apply_"))
    try:
        # working copy
        wc = work / wb_path.name
        shutil.copy2(wb_path, wc)

        # write notes by UID into the Notes column; force recalc-on-load
        wbf = openpyxl.load_workbook(wc)
        ws = wbf["collection"]
        hdr = [c.value for c in ws[1]]
        uid_i = hdr.index("UID")
        note_i = hdr.index("Notes")
        row_for_uid = {}
        for row in ws.iter_rows(min_row=2):
            u = row[uid_i].value
            if isinstance(u, (int, float)):
                row_for_uid[int(u)] = row[note_i].row
        applied_note_uids = []
        for uid, note in notes.items():
            r = row_for_uid.get(uid)
            if r:
                ws.cell(row=r, column=note_i + 1, value=note)
                applied_note_uids.append(uid)
        wbf.calculation.fullCalcOnLoad = True
        wbf.save(wc)
        wbf.close()

        # LibreOffice recalc
        recdir = work / "recalc"
        recdir.mkdir()
        rec = _recalc(soffice, wc, recdir)

        # verify
        nrows, nkeys = _verify_readback(rec, notes, applied_note_uids)
        _check_no_regression(wb_path, rec, work)

        # optimistic lock
        if wb_path.stat().st_mtime_ns != start_mtime:
            return {"ok": False, "error": "workbook changed during apply — aborted, nothing written."}, 409

        # backup then atomic replace
        backup_dir = config.DATA_DIR / "Backups"
        backup_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y.%m.%d %H%M")
        backup = backup_dir / f"{stamp} - Orchids - pre-apply.xlsx"
        shutil.copy2(wb_path, backup)
        tmp_dest = wb_path.with_suffix(".xlsx.new")
        shutil.copy2(rec, tmp_dest)
        tmp_dest.replace(wb_path)

        # covers + bookkeeping
        applied_covers = _merge_covers(covers)
        summary = f"{len(applied_note_uids)} notes, {applied_covers} covers ({nkeys}/{nrows} keys)"
        _log_and_clear(summary, applied_note_uids, list(covers.keys()))

        return {
            "ok": True,
            "applied_notes": len(applied_note_uids),
            "applied_covers": applied_covers,
            "backup": str(backup),
            "note": "Run build_collection.py to refresh the viewers (v1.2 / issue #6).",
        }, 200
    finally:
        shutil.rmtree(work, ignore_errors=True)
